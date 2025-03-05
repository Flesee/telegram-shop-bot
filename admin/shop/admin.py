from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from .models import Category, Product, Order, OrderItem, User, CartItem, FAQ, Mailing
from django.db import models


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name_with_level', 'parent', 'slug', 'created_at', 'updated_at']
    list_filter = ['parent']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name']
    date_hierarchy = 'created_at'

    def name_with_level(self, obj):
        if not obj.parent:
            return format_html('<strong>{}</strong>', obj.name)
        return format_html('→ {}', obj.name)

    name_with_level.short_description = 'Название'

    def get_queryset(self, request):
        qs = super().get_queryset(request)

        main_categories = qs.filter(parent__isnull=True).order_by('name')
        ordered_categories = []
        for category in main_categories:
            ordered_categories.append(category.id)
            subcategories = qs.filter(parent=category).order_by('name')
            ordered_categories.extend(subcategory.id for subcategory in subcategories)
        # Возвращаем отсортированный QuerySet
        return qs.filter(id__in=ordered_categories).order_by(
            models.Case(*[models.When(id=id, then=pos) for pos, id in enumerate(ordered_categories)])
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "parent":
            # Исключаем текущую категорию из списка возможных родителей
            if request.resolver_match.kwargs.get('object_id'):
                kwargs["queryset"] = Category.objects.exclude(
                    id=request.resolver_match.kwargs.get('object_id')
                ).order_by('parent__id', 'name')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class CategoryListFilter(admin.SimpleListFilter):
    title = 'Категория'
    parameter_name = 'category'

    def lookups(self, request, model_admin):
        def format_category(category, level=0, prefix=''):
            yield (
                category.id,
                format_html(
                    '<span style="color: {}; margin-left: {}px">{}{}</span>',
                    '#2c5282' if level == 0 else '#4a5568',
                    20 * level,
                    prefix,
                    category.name
                )
            )
            children = Category.objects.filter(parent=category).order_by('name')
            last_idx = len(children) - 1
            for idx, child in enumerate(children):
                is_last = idx == last_idx
                child_prefix = prefix + ('└── ' if is_last else '├── ')
                next_prefix = prefix + ('    ' if is_last else '│   ')
                yield from format_category(child, level + 1, child_prefix)

        categories = Category.objects.filter(parent__isnull=True).order_by('name')
        return [item for category in categories for item in format_category(category)]

    def queryset(self, request, queryset):
        return queryset.filter(category_id=self.value()) if self.value() else queryset


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'price', 'available', 'created_at', 'updated_at']
    list_filter = [CategoryListFilter, 'available', 'created_at', 'updated_at']
    list_editable = ['price', 'available']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name', 'description']
    date_hierarchy = 'created_at'
    readonly_fields = ['created_at', 'updated_at']

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "category":
            # Получаем все категории
            categories = Category.objects.all()

            formatted_names = {}
            for category in categories.filter(parent__isnull=True):
                formatted_names[category.id] = category.name

                # Форматируем подкатегории с отступами
                for subcategory in categories.filter(parent=category):
                    formatted_names[subcategory.id] = f"↳ {subcategory.name} (в {category.name})"

            # Переопределяем метод __str__ для категорий в контексте этой формы
            def get_formatted_name(obj):
                return formatted_names.get(obj.id, obj.name)

            Category.__str__ = get_formatted_name

        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    raw_id_fields = ['product']
    extra = 0


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['id', 'user_id', 'username', 'full_name', 'status', 'payment_status', 'total_price', 'created_at']
    list_filter = ['status', 'payment_status', 'created_at']
    search_fields = ['user_id', 'username', 'full_name', 'phone']
    date_hierarchy = 'created_at'
    inlines = [OrderItemInline]
    readonly_fields = ['created_at', 'updated_at']
    actions = ['export_to_excel']
    fieldsets = (
        ('Информация о пользователе', {
            'fields': ('user_id', 'username', 'full_name', 'phone', 'address')
        }),
        ('Информация о заказе', {
            'fields': ('status', 'total_price', 'created_at', 'updated_at')
        }),
        ('Информация об оплате', {
            'fields': ('payment_id', 'payment_status')
        }),
    )

    def export_to_excel(self, request, queryset):
        """Экспорт выбранных заказов в Excel"""
        # Создаем новый файл Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"

        # Определяем стили
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        status_colors = {
            'new': 'FFD966',  # Желтый
            'pending': 'FFD966',  # Желтый
            'paid': '9BBB59',  # Зеленый
            'completed': '9BBB59',  # Зеленый
            'cancelled': 'FF0000'  # Красный
        }

        # Устанавливаем заголовки
        headers = [
            'ID заказа', 'ID пользователя', 'Имя пользователя', 'ФИО', 'Телефон',
            'Адрес', 'Статус заказа', 'Статус оплаты', 'Сумма (₽)', 'Дата создания',
            'Товары в заказе'
        ]

        # Применяем стили к заголовкам
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Заполняем данные
        for row, order in enumerate(queryset, 2):
            # Получаем товары заказа
            order_items = OrderItem.objects.filter(order=order).select_related('product')
            items_list = []
            total_items = 0
            for item in order_items:
                items_list.append(f"• {item.product.name} ({item.quantity} шт. × {float(item.price)} ₽)")
                total_items += item.quantity
            items_str = "\n".join(items_list)

            # Заполняем основные данные
            data = [
                order.id,
                order.user_id,
                order.username,
                order.full_name,
                order.phone,
                order.address,
                order.status,
                order.payment_status,
                float(order.total_price),
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                f"Всего товаров: {total_items} шт.\n\n{items_str}"
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border

                # Применяем специальные стили в зависимости от типа данных
                if col == 9:  # Сумма
                    cell.number_format = '#,##0.00 ₽'
                    cell.alignment = number_alignment
                elif col in [7, 8]:  # Статусы
                    status_color = status_colors.get(value.lower(), 'FFFFFF')
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 11:  # Список товаров
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    # Увеличиваем высоту строки для списка товаров
                    ws.row_dimensions[row].height = max(15 * (len(items_list) + 3), 30)
                else:
                    cell.alignment = data_alignment

        # Автоматически регулируем ширину столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
            ws.column_dimensions[column_letter].width = adjusted_width

        # Замораживаем верхнюю строку
        ws.freeze_panes = 'A2'

        # Создаем HTTP-ответ с файлом Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response

    export_to_excel.short_description = "Экспортировать выбранные заказы в Excel"


@admin.register(OrderItem)
class OrderItemAdmin(admin.ModelAdmin):
    list_display = ['order', 'product', 'price', 'quantity']
    list_filter = ['order']
    search_fields = ['order__user_id', 'product__name']
    readonly_fields = ['order', 'product', 'price', 'quantity']


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ['user_id', 'username', 'full_name', 'phone', 'created_at']
    list_filter = ['created_at']
    search_fields = ['user_id', 'username', 'full_name', 'phone']
    date_hierarchy = 'created_at'
    readonly_fields = ['created_at', 'updated_at']
    fieldsets = (
        ('Основная информация', {
            'fields': ('user_id', 'username', 'full_name')
        }),
        ('Контактная информация', {
            'fields': ('phone', 'address')
        }),
        ('Даты', {
            'fields': ('created_at', 'updated_at')
        }),
    )


class CartItemInline(admin.TabularInline):
    model = CartItem
    raw_id_fields = ['product']
    extra = 0


@admin.register(CartItem)
class CartItemAdmin(admin.ModelAdmin):
    list_display = ['user', 'product', 'quantity', 'created_at']
    list_filter = ['created_at']
    search_fields = ['user__user_id', 'user__username', 'product__name']
    date_hierarchy = 'created_at'
    readonly_fields = ['created_at', 'updated_at']


@admin.register(FAQ)
class FAQAdmin(admin.ModelAdmin):
    list_display = ['question', 'created_at', 'updated_at']
    search_fields = ['question', 'answer', 'keywords']
    date_hierarchy = 'created_at'
    readonly_fields = ['created_at', 'updated_at']
    fieldsets = (
        ('Вопрос и ответ', {
            'fields': ('question', 'answer')
        }),
        ('Дополнительная информация', {
            'fields': ('keywords', 'created_at', 'updated_at')
        }),
    )


@admin.register(Mailing)
class MailingAdmin(admin.ModelAdmin):
    list_display = ['scheduled_at', 'is_sent', 'created_at']
    list_filter = ['is_sent', 'scheduled_at', 'created_at']
    readonly_fields = ['is_sent', 'created_at', 'updated_at']
    search_fields = ['text']
    date_hierarchy = 'scheduled_at'
